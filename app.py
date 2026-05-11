from __future__ import annotations

from functools import wraps
from io import BytesIO
import secrets
from uuid import uuid4

from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

from cell_service import ProductionCellService
from opcua_server import OpcUaSimulator


app = Flask(__name__)
app.secret_key = "development-only-change-before-production-" + secrets.token_hex(16)
service = ProductionCellService()
opcua_simulator = OpcUaSimulator(service)


ROLE_ORDER = ProductionCellService.ROLE_ORDER


def current_user() -> dict | None:
    user = session.get("user")
    return user if isinstance(user, dict) else None


def has_role(role: str) -> bool:
    user = current_user()
    if not user:
        return False
    return ROLE_ORDER.get(user.get("role", ""), 0) >= ROLE_ORDER.get(role, 999)


def require_role(role: str):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user():
                if request.path.startswith("/api/"):
                    return jsonify({"error": "Login required"}), 401
                return redirect(url_for("login_page", next=request.path))
            if not has_role(role):
                if request.path.startswith("/api/"):
                    return jsonify({"error": "Insufficient role"}), 403
                return redirect(url_for("dashboard_page"))
            return fn(*args, **kwargs)

        return wrapper

    return decorator


@app.context_processor
def inject_user():
    user = current_user()
    return {"current_user": user, "has_role": has_role}


@app.get("/")
def dashboard_page():
    return render_template("dashboard.html")


@app.get("/login")
def login_page():
    return render_template("login.html", next_url=request.args.get("next", "/"))


@app.post("/login")
def login_submit():
    user = service.authenticate_user(request.form.get("username", ""), request.form.get("password", ""))
    if not user:
        return render_template("login.html", error="Feil brukernavn eller passord", next_url=request.form.get("next", "/")), 401
    session["user"] = user
    return redirect(request.form.get("next") or url_for("dashboard_page"))


@app.post("/logout")
def logout_submit():
    session.clear()
    return redirect(url_for("dashboard_page"))


@app.get("/shelves")
@require_role("operator")
def shelves_page():
    return render_template("shelves.html")


@app.get("/parts")
@require_role("innstiller")
def parts_page():
    return render_template("parts.html")


@app.get("/jobs")
@require_role("operator")
def jobs_page():
    return render_template("jobs.html")


@app.get("/cnc")
@require_role("innstiller")
def cnc_page():
    return render_template("cnc.html")


@app.get("/lift")
@require_role("innstiller")
def lift_page():
    return render_template("lift.html")


@app.get("/opcua")
@require_role("service")
def opcua_page():
    return render_template("opcua.html")


@app.get("/stats")
@require_role("operator")
def stats_page():
    return render_template("stats.html")


@app.get("/diagnostics")
@require_role("service")
def diagnostics_page():
    return render_template("diagnostics.html")


@app.get("/simulation")
@require_role("service")
def simulation_page():
    return render_template("simulation.html")


@app.get("/admin")
@require_role("administrator")
def admin_page():
    return render_template("admin.html")


@app.get("/api/state")
def api_state():
    return jsonify(service.get_state())


@app.get("/api/session")
def api_session():
    return jsonify({"user": current_user(), "roles": list(ROLE_ORDER.keys())})


@app.get("/api/admin/users")
@require_role("administrator")
def api_admin_users():
    return jsonify(service.list_users())


@app.post("/api/admin/users")
@require_role("administrator")
def api_admin_users_save():
    try:
        return jsonify(service.upsert_user(request.get_json(force=True)))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.delete("/api/admin/users/<username>")
@require_role("administrator")
def api_admin_users_delete(username: str):
    try:
        return jsonify(service.delete_user(username))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/diagnostics")
@require_role("service")
def api_diagnostics():
    return jsonify(service.diagnostics)


@app.get("/api/leanlift/shelf-layout/<shelf>")
@require_role("operator")
def api_shelf_layout(shelf: str):
    try:
        return jsonify(service.get_shelf_layout(shelf))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/leanlift/shelf-layout/<shelf>/export")
@require_role("operator")
def api_shelf_layout_export(shelf: str):
    try:
        rows = service.export_shelf_layout_rows(shelf)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Shelf {shelf}"
    sheet.append(["slot_no", "x_mm", "y_mm", "z_mm", "part_no"])
    for row in rows:
        sheet.append([row["slot_no"], row["x_mm"], row["y_mm"], row["z_mm"], row["part_no"]])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"shelf_{shelf}_layout.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/leanlift/shelf-layout/<shelf>/import")
@require_role("innstiller")
def api_shelf_layout_import(shelf: str):
    uploaded_file = request.files.get("file")
    if not uploaded_file or not uploaded_file.filename:
        return jsonify({"error": "Layout file is required"}), 400
    try:
        workbook = load_workbook(uploaded_file.stream, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or "").strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not values or all(value in {None, ""} for value in values):
                continue
            row = {header: values[index] if index < len(values) else "" for index, header in enumerate(headers)}
            rows.append(row)
        return jsonify(
            service.import_shelf_layout_rows(
                shelf,
                request.form.get("part_type_id", ""),
                rows,
                float(request.form.get("part_clearance_mm", service.settings["part_clearance_mm"])),
                float(request.form.get("wall_clearance_mm", service.settings["wall_clearance_mm"])),
            )
        )
    except (ValueError, KeyError, TypeError) as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/settings")
@require_role("innstiller")
def api_settings():
    try:
        return jsonify(service.update_settings(request.get_json(force=True)))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/shelf/configure-graphic")
@require_role("innstiller")
def api_shelf_configure_graphic():
    payload = request.get_json(force=True)
    try:
        return jsonify(
            service.configure_shelf_layout_graphic(
                str(payload.get("shelf", "")),
                payload.get("part_type_id", ""),
                int(payload.get("cols", 4)),
                int(payload.get("rows", 3)),
                float(payload.get("z_mm", 100)),
                float(payload.get("part_clearance_mm", service.settings["part_clearance_mm"])),
                float(payload.get("wall_clearance_mm", service.settings["wall_clearance_mm"])),
            )
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/shelf/slot/update")
@require_role("operator")
def api_slot_update():
    payload = request.get_json(force=True)
    try:
        return jsonify(
            service.update_slot(
                str(payload.get("shelf", "")),
                int(payload.get("slot_no", 0)),
                bool(payload.get("occupied", False)),
                payload.get("status", "empty"),
                payload.get("part_type_id"),
            )
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/shelf/status-bulk")
@require_role("operator")
def api_shelf_status_bulk():
    payload = request.get_json(force=True)
    try:
        return jsonify(
            service.set_shelf_status(
                str(payload.get("shelf", "")),
                payload.get("status", "raw"),
                bool(payload.get("include_empty", False)),
                payload.get("part_type_id"),
            )
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/parts")
@require_role("innstiller")
def api_parts():
    try:
        return jsonify(service.upsert_part_type(request.get_json(force=True)))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/nc-programs/upload")
@require_role("operator")
def api_nc_program_upload():
    uploaded_file = request.files.get("file")
    if not uploaded_file or not uploaded_file.filename:
        return jsonify({"error": "NC program file is required"}), 400
    safe_name = secure_filename(uploaded_file.filename)
    if not safe_name:
        return jsonify({"error": "Invalid filename"}), 400
    service.upload_dir.mkdir(parents=True, exist_ok=True)
    stored_path = service.upload_dir / f"{uuid4().hex}_{safe_name}"
    uploaded_file.save(stored_path)
    try:
        return jsonify(service.register_uploaded_nc_program(uploaded_file.filename, str(stored_path), request.form.get("program_name")))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/jobs")
@require_role("operator")
def api_jobs_create():
    try:
        return jsonify(service.create_job(request.get_json(force=True)))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/jobs/<job_id>/start")
@require_role("operator")
def api_jobs_start(job_id: str):
    payload = request.get_json(silent=True) or {}
    try:
        return jsonify(service.start_job(job_id, payload.get("mode", "quantity"), payload.get("quantity")))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/production/start")
@require_role("operator")
def api_production_start():
    payload = request.get_json(force=True)
    try:
        return jsonify(service.start_production(payload.get("part_type_id", ""), payload.get("mode", "quantity"), payload.get("quantity")))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/production/stop")
@require_role("operator")
def api_production_stop():
    payload = request.get_json(silent=True) or {}
    return jsonify(service.stop_production(payload.get("reason", "Stopped from HMI")))


@app.post("/api/cnc/program/select")
@require_role("innstiller")
def api_cnc_program_select():
    payload = request.get_json(force=True)
    try:
        return jsonify(service.select_cnc_program(payload.get("product_id", ""), payload.get("program", ""), payload.get("operator", "development")))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/cnc/focas")
@require_role("service")
def api_cnc_focas():
    payload = request.get_json(force=True)
    try:
        return jsonify(service.call_cnc_focas(payload.get("function_name", ""), payload.get("params", {})))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/lift/command")
@require_role("service")
def api_lift_command():
    payload = request.get_json(force=True)
    try:
        return jsonify(service.call_lift_rest(payload.get("function_name", ""), payload.get("params", {})))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/lift/request-shelf")
@require_role("operator")
def api_lift_request_shelf():
    payload = request.get_json(force=True)
    try:
        return jsonify(
            service.request_shelf(
                str(payload.get("shelf", "")),
                int(payload.get("access_point", service.lift_status["operator_access_point"])),
                payload.get("actor", "operator"),
                bool(payload.get("override", False)),
            )
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/opcua/status")
@require_role("service")
def api_opcua_status():
    return jsonify(service.get_state()["opcua"])


@app.post("/api/opcua/start")
@require_role("service")
def api_opcua_start():
    started = opcua_simulator.start()
    state = service.get_state()["opcua"]
    state["start_requested"] = started
    return jsonify(state)


@app.post("/api/opcua/signal")
@require_role("service")
def api_opcua_signal():
    payload = request.get_json(force=True)
    if "command" in payload:
        return jsonify(service.apply_sim_command(payload.get("command", ""), source="web"))
    return jsonify(service.update_machine_signals("web", payload))


@app.post("/api/simulation/start")
@require_role("service")
def api_sim_start():
    return jsonify(service.simulation_start())


@app.post("/api/simulation/pause")
@require_role("service")
def api_sim_pause():
    return jsonify(service.simulation_pause())


@app.post("/api/simulation/step")
@require_role("service")
def api_sim_step():
    return jsonify(service.simulation_step())


def start_background_services() -> None:
    opcua_simulator.start()


if __name__ == "__main__":
    start_background_services()
    app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=False)
